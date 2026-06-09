import json
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# 样式定义
header_font_white = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
stock_name_font = Font(bold=True, size=16, color="1F4E79")
stock_name_fill = PatternFill(start_color="E8F4FC", end_color="E8F4FC", fill_type="solid")
formula_title_font = Font(bold=True, size=11, color="333333")
buy_formula_font = Font(size=11, color="00B050", bold=True)
sell_formula_font = Font(size=11, color="FF0000", bold=True)
profit_formula_font = Font(size=11, color="FF6B35", bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def parse_amount(s):
    """解析金额字符串，如 '$1,568.00' -> 1568.0"""
    if s is None or s == '':
        return 0.0
    return float(s.replace('$', '').replace(',', ''))


def parse_quantity(s):
    """解析数量字符串"""
    if s is None or s == '':
        return 0.0
    return float(s.replace(',', ''))


def parse_price(s):
    """解析价格字符串，如 '$392.00' -> 392.0"""
    if s is None or s == '':
        return 0.0
    return float(s.replace('$', '').replace(',', ''))


def parse_fee(s):
    """解析手续费字符串"""
    if s is None or s == '':
        return 0.0
    val = s.replace('$', '').replace(',', '')
    return float(val) if val else 0.0


def parse_date(date_str):
    """解析日期字符串 'MM/DD/YYYY' -> 标准格式"""
    # 处理 'as of' 后缀
    date_str = date_str.split(' as of')[0].strip()
    parts = date_str.split('/')
    return f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"


def create_stock_worksheet(ws, stock_name, transactions):
    """创建单个股票持仓记录工作表，使用实际交易数据"""

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].hidden = True  # 隐藏列，用于当前持仓周期内累计卖出盈亏

    # 隐藏网格线
    ws.sheet_view.showGridLines = False

    # 第1行：股票名称标题（动态引用工作表名称）
    ws.merge_cells('A1:H1')
    ws['A1'] = '="📈 " & MID(CELL("filename",A1),FIND("]",CELL("filename",A1))+1,100) & " 持仓成本记录"'
    ws['A1'].font = stock_name_font
    ws['A1'].fill = stock_name_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 第2行：分隔行
    ws.merge_cells('A2:H2')
    ws['A2'] = '-' * 80
    ws['A2'].font = Font(color="4472C4")
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 12

    # 第3行：公式标题
    ws.merge_cells('A3:H3')
    ws['A3'] = '【成本计算公式】'
    ws['A3'].font = formula_title_font
    ws['A3'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[3].height = 20

    # 第4行：买入公式
    ws.merge_cells('A4:H4')
    ws['A4'] = '🟢 买入成本 = (原成本×原股数 + 买入金额 + 手续费) ÷ (原股数 + 买入股数)'
    ws['A4'].font = buy_formula_font
    ws['A4'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[4].height = 20

    # 第5行：卖出公式
    ws.merge_cells('A5:H5')
    ws['A5'] = '🔴 卖出成本 = (原成本×原股数 - 卖出金额 + 手续费) ÷ (原股数 - 卖出股数)'
    ws['A5'].font = sell_formula_font
    ws['A5'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[5].height = 20

    # 第6行：盈亏公式
    ws.merge_cells('A6:H6')
    ws['A6'] = '💰 盈亏金额 = (卖出价格 - 持仓成本) × 卖出股数 - 手续费'
    ws['A6'].font = profit_formula_font
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[6].height = 20

    # 第7行：分隔行
    ws.merge_cells('A7:H7')
    ws['A7'] = '-' * 80
    ws['A7'].font = Font(color="4472C4")
    ws['A7'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[7].height = 12

    # 第9行：表头
    headers = ['日期', '操作', '股数', '价格', '手续费', '盈亏金额', '持仓成本', '累计持仓']
    ws.row_dimensions[9].height = 22
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 设置数据区域对齐
    for row in range(10, 105):
        for col in range(1, 9):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

    # 设置日期格式（A列）
    date_format = 'yyyy-mm-dd'
    for row in range(10, 105):
        ws.cell(row=row, column=1).number_format = date_format

    # 添加下拉数据验证（操作列：B列）
    dv = DataValidation(type="list", formula1='"买入,卖出"', allow_blank=True)
    dv.error = '请选择"买入"或"卖出"'
    dv.errorTitle = '无效的操作'
    dv.prompt = '请从下拉列表中选择'
    dv.promptTitle = '操作类型'
    ws.add_data_validation(dv)
    dv.add('B10:B104')

    # 条件格式：有输入数据时显示边框
    border_dxf = DifferentialStyle(border=thin_border)
    ws.conditional_formatting.add('A10:H104',
        Rule(type='expression', formula=['COUNTA($A10:$E10)>0'], dxf=border_dxf))

    # 操作列颜色条件格式
    green_font = Font(color="006100", bold=True)
    red_font = Font(color="9C0006", bold=True)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_dxf = DifferentialStyle(font=green_font, fill=green_fill)
    ws.conditional_formatting.add('B10:B104',
        Rule(type='expression', formula=['B10="买入"'], dxf=green_dxf))

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_dxf = DifferentialStyle(font=red_font, fill=red_fill)
    ws.conditional_formatting.add('B10:B104',
        Rule(type='expression', formula=['B10="卖出"'], dxf=red_dxf))

    # 为所有数据行（第10-104行）添加公式
    for row in range(10, 105):
        prev_row = row - 1
        # 盈亏金额公式（F列）- 处理清仓后无持仓时卖出不计算盈亏
        cell = ws.cell(row=row, column=6)
        if row == 10:
            cell.value = f'=IF(B{row}="卖出",(D{row}-G{row})*C{row}-E{row},"")'
        else:
            cell.value = f'=IF(B{row}="卖出",IF(H{prev_row}<=0,"",(D{row}-G{prev_row})*C{row}-E{row}),"")'
        cell.protection = Protection(locked=True)
        # 持仓成本公式（G列）
        if row == 10:
            cell = ws.cell(row=row, column=7, value=f'=IF(B{row}="买入",(C{row}*D{row}+E{row})/C{row},"")')
        else:
            # 清仓后重新买入：用新买入价计算；有持仓时：加权平均；卖出时保持原成本
            cell = ws.cell(row=row, column=7, value=f'=IF(B{row}="买入",IF(H{prev_row}<=0,(C{row}*D{row}+E{row})/C{row},(G{prev_row}*H{prev_row}+C{row}*D{row}+E{row})/(H{prev_row}+C{row})),IF(B{row}="卖出",IF(H{prev_row}-C{row}<=0,G{prev_row},(G{prev_row}*H{prev_row}-C{row}*D{row}+E{row})/(H{prev_row}-C{row})),""))')
        cell.protection = Protection(locked=True)
        # 累计持仓公式（H列）- 卖出清仓时返回0（不能返回""，否则后续算术运算报错）
        if row == 10:
            cell = ws.cell(row=row, column=8, value=f'=IF(B{row}="买入",C{row},IF(B{row}="卖出",-C{row},""))')
        else:
            cell = ws.cell(row=row, column=8, value=f'=IF(B{row}="买入",H{prev_row}+C{row},IF(B{row}="卖出",MAX(0,H{prev_row}-C{row}),""))')
        cell.protection = Protection(locked=True)
        # 当前持仓周期内累计卖出盈亏公式（I列，隐藏列）- 清仓(H=0)时重置为0
        if row == 10:
            cell = ws.cell(row=row, column=9, value=f'=IF(B{row}="买入",0,IF(B{row}="卖出",IF(H{row}=0,0,F{row}),0))')
        else:
            cell = ws.cell(row=row, column=9, value=f'=IF(B{row}="",I{prev_row},IF(B{row}="买入",I{prev_row},IF(B{row}="卖出",IF(H{row}=0,0,I{prev_row}+F{row}),I{prev_row})))')
        cell.protection = Protection(locked=True)

    # 第8行：盈亏平衡点 = 持仓成本 - 当前周期累计卖出盈亏 ÷ 当前持仓
    ws.merge_cells('A8:H8')
    ws['A8'] = '=IF(COUNTA(B10:B104)=0,"","⚖️ 盈亏平衡点: $"&TEXT(LOOKUP(2,1/(B10:B104<>""),G10:G104)-LOOKUP(2,1/(B10:B104<>""),I10:I104)/LOOKUP(2,1/(B10:B104<>""),H10:H104),"#,##0.00"))'
    ws['A8'].font = Font(bold=True, size=12, color="1F4E79")
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A8'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.row_dimensions[8].height = 22
    ws['A8'].protection = Protection(locked=True)

    # 写入实际交易数据
    for i, t in enumerate(transactions):
        row = 10 + i
        ws.cell(row=row, column=1, value=parse_date(t['Date']))
        # 统一操作类型
        action = t['Action']
        if action in ('Buy',):
            ws.cell(row=row, column=2, value='买入')
        elif action in ('Sell', 'Sell Short'):
            ws.cell(row=row, column=2, value='卖出')
        else:
            continue

        ws.cell(row=row, column=3, value=parse_quantity(t['Quantity']))
        ws.cell(row=row, column=4, value=parse_price(t['Price']))
        ws.cell(row=row, column=5, value=parse_fee(t.get('Fees & Comm', '')))

    # 设置可编辑区域（A-E列）为未锁定
    for row in range(10, 105):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.protection = Protection(locked=False)

    # 启用工作表保护
    ws.protection.sheet = True
    ws.protection.password = 'stock123'


def main():
    # 读取JSON文件
    json_path = os.path.join(os.path.dirname(__file__), 'position.json')
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    transactions = data['BrokerageTransactions']

    # 过滤交易动作：只保留 Buy/Sell/Sell Short
    trade_actions = {'Buy', 'Sell', 'Sell Short'}
    trades = [t for t in transactions if t['Action'] in trade_actions and t['Symbol']]

    # 计算每个股票的净持仓
    position = {}
    for t in trades:
        symbol = t['Symbol']
        qty = parse_quantity(t['Quantity'])
        if t['Action'] == 'Buy':
            net = qty
        else:  # Sell / Sell Short
            net = -qty

        if symbol not in position:
            position[symbol] = 0.0
        position[symbol] += net

    # 只保留净持仓 > 0 的股票
    held_stocks = {sym: net for sym, net in position.items() if net > 0.001}

    print(f"当前持仓股票 ({len(held_stocks)} 只):")
    for sym, net in sorted(held_stocks.items()):
        desc = next((t['Description'] for t in trades if t['Symbol'] == sym), '')
        print(f"  {sym} ({desc}): {net:.4f} 股")

    # 创建工作簿
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 为每只持仓股票创建工作表
    for symbol in sorted(held_stocks.keys()):
        # 获取该股票的所有交易，按日期升序排列，同一天内颠倒顺序（做T需要后发生的交易排在前面）
        symbol_trades = [t for t in trades if t['Symbol'] == symbol]
        stock_trades = [
            t for _, t in sorted(
                enumerate(symbol_trades),
                key=lambda x: (parse_date(x[1]['Date']), -x[0])
            )
        ]

        # 使用股票代码(symbol)作为工作表名称
        sheet_name = symbol[:31]

        ws = wb.create_sheet(title=sheet_name)
        create_stock_worksheet(ws, symbol, stock_trades)

    # 保存文件
    output_path = r'C:\Users\LiuZhen\Desktop\code\trae\stock\stock_portfolio_tracker.xlsx'
    wb.save(output_path)
    print(f"\n股票持仓成本记录表已创建：{output_path}")


if __name__ == '__main__':
    main()